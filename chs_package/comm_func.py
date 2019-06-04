import requests
from hashlib import md5
from PIL import Image
from openpyxl import load_workbook
import json
import os
import time
from datetime import datetime
from collections import OrderedDict
from selenium.webdriver.common.alert import Alert
import logging
from logging import handlers


def Get_Param_Info(Config):

    import os
    if os.path.isfile(Config) == False:
        raise Exception("错误，全局参数配置文件不存在")

    paramInfo = {}
    for line in open(Config,"r",encoding= 'UTF-8'):
        if line != "\n" :
            info = line.strip("\n")
            # 首字符为 # ; 等符号 视为注释
            if info.strip()[0] != "#" and info.strip()[0] != ";" and info.strip()[0] != "[" :
                # print(info.strip()[0])
                info = info.split("=")
                if len(info) == 2:
                    paramName = info[0].strip()
                    paramValue = info[1].strip()
                    paramInfo[paramName]=paramValue
    return paramInfo


class Login_SAC(object):
    ''' 证券业协会远程培训处理类 '''
    def __init__(self, BASE_DIR, SCREEN_PATH, SAC_ROOT_URL, PARAM_INFO, driver):
        self.BASE_DIR = BASE_DIR
        self.SCREEN_PATH = SCREEN_PATH
        self.SAC_ROOT_URL = SAC_ROOT_URL
        self.PARAM_INFO = PARAM_INFO
        self.driver = driver

    def IsElementExistClick(self, xpath):
        try:
            self.driver.find_element_by_xpath(xpath).click()
            return True
        except:
            return False

    def GetCookies(self):
        ''' 获取本地Cookies '''
        cookiesFile = r"{0}\Cookies.txt".format(self.BASE_DIR)

        try :
            # 判断文件是否存在
            isFileExists = os.path.exists(cookiesFile)

            if isFileExists == False :

                raise Exception("错误，文件不存在")

            cookiesStr =  open(cookiesFile).read()

            cookies = json.loads(cookiesStr)
            return cookies

        except Exception as e :
            return False

    def LoginSacHome(self):

        self.driver.get(self.SAC_ROOT_URL)
        self.driver.implicitly_wait(30)
        xpath = "//td/div/input[@value='关闭']"
        isNoticeExists = self.IsElementExistClick(xpath)

        # 点击登陆窗口 弹出用户 密码 验证码 输入框
        self.driver.find_element_by_id("loginReject").click()

        self.driver.find_element_by_id("id").send_keys(self.PARAM_INFO["SAC_USER"])
        self.driver.find_element_by_id("passwdpersonal").send_keys(self.PARAM_INFO["SAC_PASSWD"])
        self.driver.find_element_by_id("authCode0").click()

        loginPageShotFileName = r"{0}\login_main.png".format(self.SCREEN_PATH)
        self.driver.save_screenshot(loginPageShotFileName)

        authCode = self.driver.find_element_by_id("lyt_img_id")
        authCodeShotFileName = r"{0}\authCode.png".format(self.SCREEN_PATH)
        # authCode.getScreenshotAs(screenShotFileName)

        left = authCode.location['x']
        top = authCode.location['y']
        right = authCode.location['x'] + authCode.size['width']
        bottom = authCode.location['y'] + authCode.size['height']
        # print(left,top,right,bottom)

        authCodePNG = Image.open(loginPageShotFileName)
        authCodePNG = authCodePNG.crop((left, top, right, bottom))
        authCodePNG.save(authCodeShotFileName)

        cjyClient = CJY_Client(self.PARAM_INFO["CJY_USER"], self.PARAM_INFO["CJY_PASSWD"], self.PARAM_INFO["CJY_APPID"] )

        authCodePNG = open(authCodeShotFileName, 'rb').read()

        # http://www.chaojiying.com/price.html 识别类型及价格
        authCodeRtnDict = cjyClient.PostPic(authCodePNG, 4004)


        # 如果返回的msg不为0 则识别失败，需要再一次获取新验证码 该处需要用while
        if authCodeRtnDict["err_no"] == 0 :
            authCodeStr = authCodeRtnDict["pic_str"]
            self.driver.find_element_by_id("authCode0").send_keys(authCodeStr)

        else :
            pass

        # 登陆学习Main页面
        xpath = "//*[@id='log1']"
        self.driver.find_element_by_xpath(xpath).click()

        self.driver.implicitly_wait(30)
        cookies = self.driver.get_cookies()

        cookiesFile = open(r"{0}\Cookies.txt".format(self.BASE_DIR), "w")
        cookiesFile.write(json.dumps(cookies))
        cookiesFile.close()

    def IsAlertExistClick(self):
        try:
            self.driver.refresh()
            return True
        except :
            alertBox = Alert(self.driver)
            time.sleep(1)
            alertMsg = alertBox.text
            alertBox.accept()
            time.sleep(1)
            return False

    def Is_Design_User(self):
        userInfoUrl = "http://training.sac.net.cn/entity/workspaceStudent/studentWorkspace_StudentInfo.action"
        mainWinHandle = self.driver.current_window_handle
        js = 'window.open();'
        self.driver.execute_script(js)
        curWinHandles = self.driver.window_handles
        for newWinHandle in curWinHandles :
            if newWinHandle != mainWinHandle :
                self.driver.switch_to.window(newWinHandle)
                self.driver.get(userInfoUrl)
                time.sleep(2)
                curUserCode = self.driver.find_element_by_xpath('//*[@class="datalist"]/tbody/tr[1]/td[2]').text
                self.driver.close()
        self.driver.switch_to.window(mainWinHandle)

        if self.PARAM_INFO["SAC_USER"] == curUserCode.strip().upper() :
            return True
        else:
            return False

    def Get_Url(self, urlStr):
        try:
            self.driver.get(urlStr)
            self.driver.implicitly_wait(30)
            time.sleep(1)
            self.IsAlertExistClick()
            self.driver.implicitly_wait(30)
            return True
        except:
            return False

    def Get_QuestionUrl(self, urlStr):
        try:
            self.driver.get(urlStr)
            self.driver.implicitly_wait(30)
            time.sleep(1)
            return True
        except:
            return False

    def Get_UserHomeUrl(self, urlStr):
        try:
            self.driver.get(urlStr)
            self.driver.implicitly_wait(30)
            time.sleep(1)
            self.driver.refresh()
            return True
        except:
            rtnAlert = self.IsAlertExistClick()
            return False

    def Wait(self, sleepSecond):
        self.driver.implicitly_wait(30)
        time.sleep(sleepSecond)

class CJY_Client(object):
    '''超级鹰图片验证码识别接口'''
    def __init__(self, username, password, soft_id):
        self.username = username
        password =  password.encode('utf8')
        self.password = md5(password).hexdigest()
        self.soft_id = soft_id
        self.base_params = {
            'user': self.username,
            'pass2': self.password,
            'softid': self.soft_id,
        }
        self.headers = {
            'Connection': 'Keep-Alive',
            'User-Agent': 'Mozilla/4.0 (compatible; MSIE 8.0; Windows NT 5.1; Trident/4.0)',
        }

    def PostPic(self, im, codetype):
        """
        im: 图片字节
        codetype: 题目类型 参考 http://www.chaojiying.com/price.html
        """
        params = {
            'codetype': codetype,
        }
        params.update(self.base_params)
        files = {'userfile': ('ccc.jpg', im)}
        r = requests.post('http://upload.chaojiying.net/Upload/Processing.php', data=params, files=files, headers=self.headers)
        return r.json()

    def ReportError(self, im_id):
        """
        im_id:报错题目的图片ID
        """
        params = {
            'id': im_id,
        }
        params.update(self.base_params)
        r = requests.post('http://upload.chaojiying.net/Upload/ReportError.php', data=params, headers=self.headers)
        return r.json()


class Log_Processing(object):

    def __init__(self):
        pass

    def debug(self):
        pass

    def errinfo(self):
        pass

    def tmpinfo(self):
        pass


class Proc_Excel(object):

    def __init__(self, xlsFile):
        self.xlsFile = xlsFile
        self.xlsWBook = load_workbook(xlsFile)

    def Search_Xls_Content(self, xlsSheetName, seachStr):
        ''' 搜索excel内容 全表扫描模式 '''
        xlsWSheet = self.xlsWBook[xlsSheetName]

        startRow = 2  # 起始行号
        xlsRowList = []
        colCnt = len(list(xlsWSheet.columns)) # 列
        rowCnt = len(list(xlsWSheet.rows))    # 行
        xlsWSheet = list(xlsWSheet.rows)
        xSn = 0
        # print(colCnt,rowCnt,xlsWSheet)
        msg = OrderedDict()
        for rowv in range(1, rowCnt):
            '''行循环'''
            for colv in range(0, colCnt):
                # print(seachStr, xlsWSheet[rowv][colv].value)
                if xlsWSheet[rowv][colv].value == seachStr :
                    msg["rowv"] = rowv
                    msg["colv"] = colv
                    msg["msg"] = True
                    return msg

        msg["rowv"] = 0
        msg["colv"] = 0
        msg["msg"] = False
        return msg

    def Write_Index_Form(self, strNo, strName):
        ''' 写入课程记录 '''
        xlsSheetName = "index"
        xlsWSheet = self.xlsWBook[xlsSheetName]
        xlsWSheet.append([strNo, strName])

        ''' 创建一个课程的sheet '''
        if strNo not in self.xlsWBook.sheetnames :
            noSheet = self.xlsWBook.create_sheet(title=strNo)
            subExcelSheelName = [
                "题目类型",
                "题目项",
                "A",
                "B",
                "C",
                "D",
                "E",
                "答案",
            ]
            noSheet.append(subExcelSheelName)

        self.xlsWBook.save(self.xlsFile)

    def Insert_Sub_Xls(self, xlsSheetName, testQuestionAnsInfo):
        xlsSheet = self.xlsWBook[xlsSheetName]
        xlsSheet.append(testQuestionAnsInfo)
        self.xlsWBook.save(self.xlsFile)

    def Get_ExcludeList(self):
        # 读取Excek文件
        xlsWSheet = self.xlsWBook["index"]
        startRow = 2  # 起始行号
        xlsRowList = []

        colCnt = len(list(xlsWSheet.columns))
        rowCnt = len(list(xlsWSheet.rows))
        xlsWSheet = list(xlsWSheet.rows)
        xSn = 0
        for rowv in range(1, rowCnt):
            xSn = xSn + 1
            xValue = xlsWSheet[rowv][0].value
            flagValue = xlsWSheet[rowv][5].value
            # print(xValue, flagValue)
            if flagValue == "Y" :
                xlsRowList.append(xValue)
            # colDict={}
            # for colv in range(0, colCnt):
            #     if xlsWSheet[rowv][colv].value == None:
            #         valueX = ""
            #     else:
            #         valueX = xlsWSheet[rowv][colv].value
            #
            #         # 如果Excel里的变量为字符型，则去前后空格，减免出错几率
            #         if isinstance(valueX,str) == True:
            #             valueX = valueX.strip()
            #
            #     colDict[xlsWSheet[0][colv].value] = valueX
            #     colDict["xSn"] = xSn
            # if int(colDict["处理标志"][0]) == 0:
            #     xlsRowList.append(colDict)

        return xlsRowList

    def Get_QuestionAnsList(self, xlsSheetName):
        '''取题库中的题目及答案列表'''
        # 读取Excek文件
        xlsWSheet = self.xlsWBook[xlsSheetName]
        startRow = 2  # 起始行号
        xlsRowList = []

        colCnt = len(list(xlsWSheet.columns))
        rowCnt = len(list(xlsWSheet.rows))
        xlsWSheet = list(xlsWSheet.rows)
        xSn = 0
        for rowv in range(1, rowCnt):
            xSn = xSn + 1
            questionStr = xlsWSheet[rowv][1].value
            ansStr = xlsWSheet[rowv][7].value
            # print(xValue, flagValue)
            if ansStr != "" and ansStr != None:
                questionDict = OrderedDict()
                questionDict["QUESTION_STR"] = questionStr
                questionDict["ANS_STR"] = ansStr
                questionDictStr = str(questionDict)
                questionDictNew = eval(questionDictStr)
                xlsRowList.append(questionDictNew)

        return xlsRowList

    def Get_SingAns(self, xlsSheetName, questionStr):
        '''取单一题目的答案'''
        # 读取Excek文件
        xlsWSheet = self.xlsWBook[xlsSheetName]
        startRow = 2  # 起始行号
        colCnt = len(list(xlsWSheet.columns))
        rowCnt = len(list(xlsWSheet.rows))
        xlsWSheet = list(xlsWSheet.rows)
        xSn = 0
        for rowv in range(1, rowCnt):
            xSn = xSn + 1

            ansStr = xlsWSheet[rowv][7].value
            # print(xValue, flagValue)
            if ansStr != "" and ansStr != None and questionStr == xlsWSheet[rowv][1].value:
                return ansStr

import re
def Question_Comparison(examDivElems, xlsQuestionAnsList):
    pageTestQuestionList = []
    for examDivElem in examDivElems:
        # testQuestionAnsInfo = []
        testQuestion = examDivElem.find_element_by_class_name("test_item_tit").text
        testQuestion = re.sub("\d \. ", "", testQuestion)
        pageTestQuestionList.append(testQuestion)

    # 判断页面上的题目是不是都包含在已获取的题库中 如果不包含 需要重新刷新一次获取
    i = 0
    for pageTestQuestion in pageTestQuestionList:

        for xlsQuestion in xlsQuestionAnsList:

            if xlsQuestion["QUESTION_STR"] == pageTestQuestion:
                i += 1
    return i

class Logger(object):
    level_relations = {
        'debug':logging.DEBUG,
        'info':logging.INFO,
        'warning':logging.WARNING,
        'error':logging.ERROR,
        'crit':logging.CRITICAL
    }#日志级别关系映射

    def __init__(self,filename,level='info',when='D',backCount=3,fmt='%(asctime)s - %(pathname)s[line:%(lineno)d] - %(levelname)s: %(message)s'):
        # log.logger.debug('debug')
        # log.logger.info('info')
        # log.logger.warning('警告')
        # log.logger.error('报错')
        # log.logger.critical('严重')
        # Logger('error.log', level='error').logger.error('error')

        fmt = '%(asctime)s - %(levelname)s: %(message)s'
        self.logger = logging.getLogger(filename)
        format_str = logging.Formatter(fmt)#设置日志格式
        self.logger.setLevel(self.level_relations.get(level))#设置日志级别
        sh = logging.StreamHandler()#往屏幕上输出
        sh.setFormatter(format_str) #设置屏幕上显示的格式
        th = handlers.TimedRotatingFileHandler(filename=filename,when=when,backupCount=backCount,encoding='utf-8')#往文件里写入#指定间隔时间自动生成文件的处理器
        #实例化TimedRotatingFileHandler
        #interval是时间间隔，backupCount是备份文件的个数，如果超过这个个数，就会自动删除，when是间隔的时间单位，单位有以下几种：
        # S 秒
        # M 分
        # H 小时、
        # D 天、
        # W 每星期（interval==0时代表星期一）
        # midnight 每天凌晨
        th.setFormatter(format_str)#设置文件里写入的格式
        self.logger.addHandler(sh) #把对象加到logger里
        self.logger.addHandler(th)
